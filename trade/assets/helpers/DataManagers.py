# Description: Module for managing data for options

#To-do:
# Clean up code
# Write documentation
# Write tests
# Create Abstract Base Class for OptionDataManager - DataManager

from dotenv import load_dotenv
import os
import sys
load_dotenv()
sys.path.append(os.environ.get('DBASE_DIR'))
sys.path.append(os.environ.get('WORK_DIR'))
import logging
from openpyxl import load_workbook
from datetime import datetime, date
import pandas as pd
import threading
from pathos.multiprocessing import ProcessingPool as Pool
from concurrent.futures import ProcessPoolExecutor, as_completed
import concurrent.futures
from trade.assets.Stock import Stock
from trade.helpers.helper import generate_option_tick
from trade.assets.rates import get_risk_free_rate_helper
from trade.helpers.helper import IV_handler, time_distance_helper, binomial_implied_vol, wait_for_response
from trade.helpers.helper import extract_numeric_value, change_to_last_busday
from trade.helpers.Logging import setup_logger
from trade.assets.Calculate import Calculate
from trade.helpers.Context import Context
from dbase.DataAPI.ThetaData import retrieve_ohlc, retrieve_quote_rt, retrieve_eod_ohlc, resample, retrieve_quote
from dbase.DataAPI.Organizers import generate_optionData_to_save, Calc_Risks
from dbase.database.SQLHelpers import store_SQL_data_Insert_Ignore, query_database, dynamic_batch_update
from trade.helpers.decorators import log_error, log_error_with_stack

OptDataManagerLogger = setup_logger('OptionDataManager_Module', stream_log_level=logging.CRITICAL)
logger = setup_logger('OptionDataManager.py', stream_log_level=logging.CRITICAL)


def write_to_excel(excel_path, sheetname, data):
    wkbook = load_workbook(excel_path)
    sheetname_exists = sheetname in wkbook.sheetnames

    if not isinstance(data, pd.DataFrame):
        raise ValueError("Data must be a pandas DataFrame")
    
    if sheetname_exists:
        existing_data = pd.read_excel(excel_path, sheet_name = sheetname)
        combined_data = pd.concat([existing_data, data], ignore_index = True)
    else:
        combined_data = data

    
    with pd.ExcelWriter(excel_path, engine='openpyxl', mode = 'a', if_sheet_exists='replace') as writer:
        combined_data.drop_duplicates(inplace = True)
        combined_data.to_excel(writer, sheet_name = sheetname, index = False)


class OptionDataManager:
    """
    Class for querying database & ensuring data integrity for options
    """
    tables = {'EOD': 'TEMP_OPTIONS_EOD', 'INTRA': 'TEMP_OPTIONS_INTRA'}
    tmfrms = [ 'h', 'd', 'w', 'M', 'q', 'y']

    def __init__(self,
                symbol: str,
                exp: str | datetime | date,
                right: str,
                strike: float,
                default_fill: str = 'midpoint') -> None:
        """
        Returns an object for querying data

        Params:
        symbol: Underlier symbol
        exp: expiration
        right: Put(P) or Call (C)
        strike: Option Strike
        default_fill: How to fill zero values for close. 'midpoint' or 'weighted_midpoint'
        """
        assert isinstance(strike, float), f"Strike has to be type float, recieved {type(strike)}"
        if default_fill not in ['midpoint', 'weighted_midpoint', None]:
            raise ValueError("Expected default_fill to be one of: 'midpoint', 'weighted_midpoint', None ")
        self.exp = exp
        self.symbol = symbol
        self.right = right.upper()
        self.strike = strike
        self.default_fill = default_fill
        self.opttick = generate_option_tick(symbol, right, exp, strike)
        self.Stock = Stock(symbol, run_chain = False)

    @log_error_with_stack(OptDataManagerLogger, False)
    def get_timeseries(self, start: str, end: str, interval: str, type_ = 'spot', model = 'bs', **kwargs):

        """
        Method Res
        """
        
        type_ = type_.lower()
        assert type_ in ['spot', 'vol', 'vega', 'vanna', 'volga', 'delta', 'gamma', 'theta', 'rho', 'greeks', 'greek'], f'expected "spot", "vol", "vega", "vanna", "volga", "delta", "gamma", "theta", "rho", "greeks", "greek" for type_, got {type_}'
        range_filters, flag, data = self.__verify_data_completeness(interval, start, end)
        self.data = data


        UseTable = self.tables[flag]
        gen_dataBool = False
        if range_filters != []:
            gen_dataBool = True
        
        elif range_filters == [] and len(data) == 0:
            gen_dataBool = True
    
        if gen_dataBool:   
            unProcessedData = self.__generate_optionData_to_save(range_filters, end, start, flag, **kwargs)
            if unProcessedData.__class__.__name__ == 'DataFrame':
                thread_data = unProcessedData.copy()
                processing_thread = threading.Thread(target = self.__process_data, args = (thread_data, flag, start, end,), name = 'SaveDataProcess')
                processing_thread.start()
                self.processing_thread = processing_thread

                # ## Keep only columns in unProcessedData
                unProcessedData.drop(columns = ['Datetime'], inplace = True)
                unProcessedData.reset_index(inplace = True)
                unProcessedData.columns = [x.lower() for x in unProcessedData.columns]
                cols = unProcessedData.columns
                data.columns = [x.lower() for x in data.columns]
                # print('In concat process')
                unProcessedData = pd.concat([data,unProcessedData ])
            else:
                OptDataManagerLogger.info(f"Data for {self.opttick} unavailable for query. It is not a dataframe, type: {type(unProcessedData)}")
                unProcessedData = data
                

        else:
            unProcessedData = data
        
        if type_ == 'spot':
            if self.default_fill is None:
                column_agg = {'Open': 'first', 'High': 'max', 'Low': 'min', 'Close': 'last', 'Volume': 'sum'}
            else:
                metric2 = self.default_fill
                column_agg = {'Open': 'first', 'High': 'max', 'Low': 'min', 'Close': 'last', metric2.capitalize(): 'last','Volume': 'sum'}
            organized_data = self.__spot_data_organizer_handler(unProcessedData)


        ## We're going to use the same function for vol, vega, vanna, volga, delta, gamma, theta, rho, greeks, greek
        ## Because greeks need vol to calculate. We will just change column_agg    
        elif type_ == 'vol' or type_ in ['vega', 'vanna', 'volga', 'delta', 'gamma', 'theta', 'rho', 'greeks', 'greek']:
            organized_data = self.__vol_data_organize_handler(unProcessedData, flag, model = model)

            # print("Organized Data")
            # print(organized_data.head())
            if model == 'bs' or model == 'bsm':
                metric = 'bs_iv'
            elif model == 'bt' or model == 'binomial':
                metric = 'binomial_iv'
            usePrice = metric.capitalize() if not self.default_fill else f"{self.default_fill}_{metric}".capitalize()
            column_agg = {usePrice:'last', metric.capitalize(): 'last'}
        
            if type_ == 'vol':
                pass

            else:
                unProcessedData = unProcessedData.set_index('datetime')
                volColumns = [x.lower() for x in column_agg.keys()]
                unProcessedData[volColumns] = organized_data[column_agg.keys()]
                unProcessedData.reset_index(inplace = True)
                organized_data = self.__greek_data_organizer_handler(unProcessedData, flag, type_)
                column_agg = dict(zip(organized_data.columns, ['last']*len(organized_data.columns)))

        return_verified_data = self.__verify_return_data_integrity(organized_data, flag)
        resampled = self.__resample(return_verified_data, interval,column_agg)
        if flag == 'INTRA':
            self.__save_intra_to_update(start, end)
        
        return resampled[(resampled.index >= start) & (resampled.index <= end)]
        
    @log_error_with_stack(OptDataManagerLogger, False)
    def __save_intra_to_update(self, start, end):
        data = pd.DataFrame({'symbol': [self.symbol], 'optiontick': [self.opttick], 'exp': [self.exp], 'right': [self.right],'default_fill': [self.default_fill], 'strike': [self.strike], 'start': [start], 'end': [end]})
        path  = f'{os.environ["JOURNAL_PATH"]}/Algo/InputFiles/IntraSaveLog.xlsx'
        write_to_excel(path, datetime.today().strftime('%Y-%m-%d'), data)



    def get_spot(self, type_ = 'close', query_date = datetime.now(), **kwargs):
        genDataKwargs = {'print_url': False}
        
        ## Check if query_date is before market open (9:30) if date is today
        # print('Spot Bitch')
        # print(f'Date Bool: {(pd.to_datetime(query_date).date() == datetime.now().date())}')
        # print(f'Time Comp Bool: {(datetime.now().time() < pd.Timestamp("9:30").time())}')


        if (pd.to_datetime(query_date).date() == datetime.now().date()):
            if (datetime.now().time() < pd.Timestamp('9:30').time()):
                query_date = change_to_last_busday(query_date).strftime('%Y-%m-%d %H:%M:%S')
            else:
                query_date = datetime.now().strftime('%Y-%m-%d')

        end_date = change_to_last_busday(datetime.today().strftime('%Y-%m-%d %H:%M:%S'))
        
        if type_ == 'close':
            if len(kwargs) > 0:
                for kwarg in genDataKwargs.keys():
                    try:
                        ## Getting all kwargs that are not in genDataKwargs
                        genDataKwargs[kwarg] = kwargs.pop(kwarg)

                    except:
                        pass
            

            ## Try intraday first
            data = retrieve_ohlc(symbol = self.symbol, end_date =query_date, 
                                exp = self.exp, right = self.right, start_date =query_date, 
                                strike = self.strike, **genDataKwargs)
                

            
            ## If we have no intraday data, go to EOD
            if data is None:
                data = retrieve_eod_ohlc(symbol = self.symbol, end_date =query_date, 
                                exp = self.exp, right = self.right, start_date =query_date, 
                                strike = self.strike, **genDataKwargs)
                
                ## Handling for unavailable data in Intraday and EOD
                if data is None:
                    OptDataManagerLogger.info(f"{self.opttick} doesn't have intraday data for {query_date}")
                    return {datetime.now().strftime('%Y-%m-%d %H:%M:%S'): "DATA UNAVAILABLE"}
                

            
            data = data[data.index.date == pd.to_datetime(query_date).date()]
            return self.__spot_data_organizer_handler(data, ts = False)
        elif type_ == 'quote':
            try:

                ## If query_date is today, we can get real time data
                if pd.to_datetime(query_date).date() == datetime.now().date():

                    data = retrieve_quote_rt(symbol = self.symbol, end_date = query_date, 
                                        exp = self.exp, right = self.right, start_date = query_date, 
                                        strike = self.strike)

                
                ## If query_date is not today, we can only get EOD data
                else:
                    data = retrieve_quote(symbol = self.symbol, end_date = query_date, 
                                        exp = self.exp, right = self.right, start_date = query_date, 
                                        strike = self.strike)
                
                if data is None:
                    OptDataManagerLogger.info(f"{self.opttick} doesn't quote have data for {query_date}")
                    return {datetime.now().strftime('%Y-%m-%d %H:%M:%S'): "DATA UNAVAILABLE"}
                
                data = data[data.index.date == pd.to_datetime(query_date).date()].tail(1)

                return data[['Bid', 'Ask', 'Midpoint', 'Weighted_midpoint']]
            except Exception as e:
                return {datetime.now().strftime('%Y-%m-%d %H:%M:%S'): "DATA UNAVAILABLE"}
        else:
            raise ValueError(f"Expected 'close' or 'quote' for type_, got {type_}")
            
        # return data


    def get_vol(self, type_ = 'close', query_date = datetime.now()):
        
        ## Check if query_date is before market open (9:30) if date is today
        # print(f'Date Bool: {(pd.to_datetime(query_date).date() == datetime.now().date())}')
        # print(f'Time Comp Bool: {(pd.to_datetime(query_date).time() < pd.Timestamp("9:30").time())}')
        # print(query_date)

        if (pd.to_datetime(query_date).date() == datetime.now().date()):
            if (pd.to_datetime(query_date).time() < pd.Timestamp('9:30').time()):
                query_date = change_to_last_busday(query_date).strftime('%Y-%m-%d %H:%M:%S')
            else:
                query_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        if type_ == 'close':

            ## Try intraday first
            data = retrieve_ohlc(symbol = self.symbol, end_date = query_date, 
                            exp = self.exp, right = self.right, start_date = query_date, 
                            strike = self.strike)
            
            ## Switch to EOD if no intraday data
            if data is None:
                OptDataManagerLogger.info(f"{self.opttick} doesn't have intraday data for {query_date}")
                OptDataManagerLogger.info(f"OptionDataManager Switching to EOD data")
                data = retrieve_eod_ohlc(symbol = self.symbol, end_date = query_date, 
                            exp = self.exp, right = self.right, start_date =query_date, 
                            strike = self.strike)
                
                ## Return unavailable data if no EOD nor Intraday data
                if data is None:
                    OptDataManagerLogger.info(f"{self.opttick} doesn't have EOD nor Intraday data for {query_date}")
                    return {query_date: "DATA UNAVAILABLE"}
                
                data = data.tail(1)

            if self.default_fill:
                price_cols = ['Close', self.default_fill.capitalize()]
            else:
                price_cols = ['Close']
        elif type_ == 'quote':
            data = self.get_spot(type_ = 'quote', query_date = query_date)
            price_cols = ['Midpoint', 'Weighted_midpoint']

        else:
            raise ValueError(f"Expected 'close' or 'quote' for type_, got {type_}")

        
        
 
        with Context(end_date = query_date, start_date = query_date, timeframe = 'day', timewidth = '1'):
            stk = Stock(self.symbol, run_chain = False)
            y = stk.div_yield()
            r = stk.rf_rate
            s0 = list(stk.spot().values())[0]

        for p in price_cols:
            if len(data) == 0 or isinstance(data, dict):
                return {datetime.now(): "DATA UNAVAILABLE"}
            data[f'{p}_Vol'] = data.apply(lambda x:IV_handler(
                                                price = x[p],
                                                S = s0,
                                                K = self.strike,
                                                t = time_distance_helper(exp = self.exp, strt = query_date),
                                                r = r,
                                                q = y,
                                                flag = self.right.lower()), axis = 1)
        price = data[price_cols + [f'{p}_Vol' for p in price_cols]]

        return price


    def __init_query(self, timeAggType:str = 'EOD'):
        
        UseTable = self.tables[timeAggType]
        self.query = f"""SELECT *
        FROM SECURITIES_MASTER.{UseTable}
        WHERE OPTIONTICK = '{self.opttick}'
        """
        database_data = query_database('securities_master', UseTable, self.query)
        self.database_data = database_data
        return database_data

    
    def __verify_data_completeness(self, interval, start, end):
        ivl_str, ivl_int = extract_numeric_value(interval)
        range_filters = []
        end = change_to_last_busday(end).strftime('%Y-%m-%d')

     

           
           ## STEP 1: RETRIEVE AVAILABLE DATA IN DATABASE 
        if ivl_str.lower() in ['d', 'w','q','y'] or ivl_str == 'M':
            ## EOD OHLCV
            busStart = start
            busEnd = end 
            timeAggType = 'EOD'
            available_data = self.__init_query(timeAggType) 
            pandas_dt_alias = 'B'
            date_range = pd.date_range(busStart, busEnd, freq = pandas_dt_alias)
        
        elif ivl_str == 'm':
            raise AttributeError("Minute data currently unavailable, please go higher")

        elif ivl_str == 'h':
            ## Intraday OHLCV
            busStart = start + ' 9:30'
            busEnd = end + ' 16:00'
            timeAggType = 'INTRA'
            available_data = self.__init_query(timeAggType) 
            pandas_dt_alias = 'H'
            date_range = pd.date_range(busStart, busEnd, freq = pandas_dt_alias)
            date_range = date_range[(date_range.time >= pd.Timestamp('9:30').time()) & (date_range.time <= pd.Timestamp('16:00').time()) & (date_range.weekday <= 5)]
            self.data_range = date_range

        ##Resample other timeframes
        else:
            raise TypeError(f"Invalid timframe. Available timeframes are: {', '.join(self.tmfrms)}. FYI: m: minute, M: month")
        
        if len(available_data) > 0:
            if timeAggType == 'INTRA':
                unavailable_dates = sorted(set(date_range) - set(pd.to_datetime(available_data.Datetime)))
            elif timeAggType == 'EOD':
                unavailable_dates = sorted(set(date_range) - set(pd.to_datetime(available_data.Datetime)))

            range_filters = list(unavailable_dates)
        
        else:
            OptDataManagerLogger.info(f"No data available for {self.opttick}. Downloading data")

        
        return range_filters, timeAggType, available_data
    
    def __generate_optionData_to_save(self, range_filters, end, start,timeAggFlag, **kwargs):
        # print("Generating Option Data to Save for this range:", range_filters)
        
        genDataKwargs = {'print_url': False}
        if len(kwargs) > 0:
            for kwarg in genDataKwargs.keys():
                try:
                    genDataKwargs[kwarg] = kwargs.pop(kwarg)

                except:
                    pass
        data = generate_optionData_to_save(self.symbol, end, self.exp,self.right, start, self.strike, range_filters = range_filters, timeAggFlag=timeAggFlag, **genDataKwargs)
        OptDataManagerLogger.info(f"Successfully downloaded unprocessed data for {self.opttick}.")
        return data
    

    def __process_data(self, data, timeAggType, start, end) -> None:

        processed_data =  Calc_Risks(data, timeAggType, self.symbol, end,self.exp, self.right, start, self.strike,)
        self.processed_data = processed_data
        OptDataManagerLogger.info(f"Successfully processed data for {self.opttick}")
        self.__save_data(processed_data, timeAggType)


    @log_error_with_stack(OptDataManagerLogger, False)
    def __save_data(self, data, timeAggType) -> None:
        OptDataManagerLogger.info(f"OptionDataManager saving data for {self.opttick}")
        UseTable = self.tables[timeAggType]
        data.drop_duplicates(inplace = True)
        store_SQL_data_Insert_Ignore('SECURITIES_MASTER',UseTable, data)
        print(f"Successfully saved {self.opttick}", end = '\r')
        self.__update_last_saved_column(UseTable)
        
    
    
    def __update_last_saved_column(self, UseTable):
        dynamic_batch_update('SECURITIES_MASTER', UseTable, {'last_updated': datetime.now()}, {'OPTIONTICK': self.opttick})
        print('Successfully updated column', end = '\r')



    @log_error_with_stack(OptDataManagerLogger, False)
    def __verify_return_data_integrity(self, data, timeAggType):

        ## Check for missing dates
        if timeAggType == 'INTRA':
            date_range = pd.date_range(data.index.min(), data.index.max(), freq = 'BH')
            date_range= date_range[(date_range.time >= pd.Timestamp('9:30').time()) & (date_range.time <= pd.Timestamp('16:00').time()) & (date_range.weekday <= 5)]
        else:
            date_range = pd.date_range(data.index.min(), data.index.max(), freq = 'B')
        
        missing_dates =set(date_range) - set(data.index)
        if len(missing_dates) > 0:
            OptDataManagerLogger.info(f"{self.opttick} is missing the following dates {missing_dates} for {timeAggType} agg")
        
        if data.reset_index().duplicated().sum() > 0:
            name = f"{self.opttick}_{datetime.today().strftime('%Y%m%d_%H:%M:%S')}"
            OptDataManagerLogger.info(f"{self.opttick} is has duplicates for {timeAggType} agg")
            OptDataManagerLogger.info(f"{self.opttick} data saved in ./DataForLogs/{name}.csv")
            if not os.path.exists('DataForLogs'):
                os.mkdir('DataForLogs')
            data.to_csv(f'./DataForLogs/{name}.csv', index = True)
        
        return data

    
    def __spot_data_organizer_handler(self, data, metric = 'close', ts = True):
        data.columns = [x.lower() for x in data.columns]
        # print(f"Full data size: {len(data)}")
        # print(f"NA Size: {len(data[data.isna().any(axis = 1)])}")
        if ts:
            if self.default_fill is None:
                metric = metric
                data = data[['datetime','open', 'high', 'low', metric, 'volume']]
            else:
                metric2 = self.default_fill
                data = data[['datetime','open', 'high', 'low', metric, metric2, 'volume']]

            data.columns = [x.capitalize() for x in data.columns]
            data.set_index('Datetime', inplace = True)
            return data
        else:
            spot = data['close'].tail(1).to_dict()
            return spot
    
    
    def __vol_data_organize_handler(self, data, timeAgg, ts = True, model = 'bs'):

        data.columns = [x.lower() for x in data.columns]
        
        ## Setting default_fill accordingly
        price = self.default_fill if self.default_fill else 'close'

        ## Keeping track of calculated data vs non calculated data
        ## nonNaData is data that has vol from db while NaData is data that needs to be calculated
        nonNaData = data[~data[['bs_iv', 'binomial_iv']].isna().any(axis = 1)]
        NaData = data[data[['bs_iv', 'binomial_iv']].isna().any(axis = 1)]

        ## Deciding which vol to calculate
        calc_vol = False
        if model == 'bs' or model == 'bsm':
            metric = 'bs_iv'
        elif model == 'bt' or model == 'binomial':
            metric = 'binomial_iv'
        else:
            raise ValueError("model arg expected 'bs' or 'bsm' for black scholes OR 'bt'/'binomial' for binomial tree. Got {model}".format(model = model))
        
        ## If we have NaN values for vol, this means we need to calculate vol
        if len(NaData) > 0 :
            calc_vol = True
        
        ## Initiating the columns to return for vol data
        if self.default_fill:
            colName = f"{self.default_fill}_{metric}"
            useCols = ['datetime',colName, metric]
        else:
            colName = metric
            useCols = ['datetime',colName]




        if calc_vol:
            ## IF WE HAVE TO CALC VOL FIRST CHECK IF THREAD IS STILL RUNNING
            isThreadRunning = self.processing_thread.is_alive()
            thread_func = lambda : not self.processing_thread.is_alive()
            
            ## Check if the database saving thread is running
            if isThreadRunning:

                ## Wait for response from database
                wait_for_response(15, thread_func, 1)
                isThreadRunning = self.processing_thread.is_alive()

            ## If no response, calculate vol
            if isThreadRunning:
                print("OptionDataManager calculating vol. Database unavailable", end = '\r')
                OptDataManagerLogger.info(f"Data for {self.opttick} unavailable for query, proceeding to calculate vol")
                if metric == 'bs_iv':
                    NaData[colName] = self.__bs_vol(NaData, price = price)
                    if self.default_fill:
                        close_vol= self.__bs_vol(NaData, price = 'close')
                        NaData[metric] = close_vol
                
                else:
                    NaData[colName] = self.__binomial_vol(NaData, price = price)
                    if self.default_fill:
                        close_vol= self.__binomial_vol(NaData, price = 'close')
                        NaData[metric] = close_vol
                # print(NaData[[colName, 'binomial_iv']])
                full_data = pd.concat([NaData, nonNaData])
                # print(f"Na Data: {len(NaData)}")
                # print(f"Non NA Data: {len(nonNaData)}")
                # print(f"Concat Data: {len(full_data)}")
                data = full_data[useCols]

            ## If response, query database instead    
            else:
                ## If we query database, we do not need to concat. The database has everything
                # print('Good to query database')
                if timeAgg == 'INTRA':
                    table = 'temp_options_intra'
                elif timeAgg == 'EOD':
                    table = 'temp_options_eod'
                else:
                    raise ValueError('Unknown timeAgg arguement. Recieved', timeAgg)
                data = query_database('securities_master', table, self.query)
                data.columns = [x.lower() for x in data.columns]
                data = data[useCols]

        else:
            logger.info(f'OptionDataManager Using available data for {self.opttick}')
            print("OptionDataManager Using available data", end = '\r')
            data = data[useCols]


        data.columns = [x.capitalize() for x in data.columns]
        data.set_index('Datetime', inplace = True)
        # print(data)
        return data

    def __resample(self, data, interval, columns_agg):
            return resample(data, interval, columns_agg)
    
    def __binomial_vol(self, data, price) -> pd.Series:
        result = data.apply(lambda x: binomial_implied_vol(
                                                            price = x[price],
                                                            S = x['underlier_price'],
                                                            K = x['strike'],
                                                            T = x['expiration'],
                                                            r = x['rf_rate'],
                                                            dividend_yield = x['dividend'],
                                                            option_type = x['put/call'].lower(),
                                                            pricing_date = x['datetime']
    ), axis = 1)

        return result

    def __bs_vol(self, data, price) -> pd.Series:
        result = data.apply(lambda x: IV_handler(
                                                price = x[price],
                                                S = x['underlier_price'],
                                                K = x['strike'],
                                                t = time_distance_helper(exp = x['expiration'], strt = x['datetime']),
                                                r = x['rf_rate'],
                                                q = x['dividend'],
                                                flag = x['put/call'].lower()), axis = 1)
        return result

    def __greek_data_organizer_handler(self, data, timeAgg, greek):
        returnColumns = ['datetime']
        ## Align all columns to lowercase
        if greek == 'greeks' or greek == 'greek':
            greek = ['delta', 'gamma', 'theta', 'rho', 'vega', 'vanna', 'volga']
        else:
            greek = [greek]


        data.columns = [x.lower() for x in data.columns]

        ## Setting default_fill accordingly
        price = self.default_fill if self.default_fill else 'close'

        ## Keeping track of calculated data vs non calculated data
        ## nonNaData is data that has greeks from db while NaData is data that needs to be calculated
        nonNaData = data[~data[['delta', 'gamma', 'vega']].isna().any(axis = 1)]
        NaData = data[data[['delta', 'gamma', 'vega']].isna().any(axis = 1)]


        ## If we have NaN values for greeks, this means we need to calculate greeks
        calc_greeks = False
        if len(NaData) > 0 :
            calc_greeks = True
        
        ## Initiating the columns to return for vol data
        returnColumns.extend(greek)
        if self.default_fill:
            major_vol_name = f"{self.default_fill}_bs_iv"
            returnColumns.extend([f"{self.default_fill}_{x}" for x in greek])
            
            

        # print(calc_greeks)
        if calc_greeks:
            ## IF WE HAVE TO CALC VOL FIRST CHECK IF THREAD IS STILL RUNNING
            isThreadRunning = self.processing_thread.is_alive()
            thread_func = lambda : not self.processing_thread.is_alive()
            # print('isThreadRunning', isThreadRunning)
            
            ## Check if the database saving thread is running
            if isThreadRunning:

                # Wait for response from database
                wait_for_response(10, thread_func, 1)
                isThreadRunning = self.processing_thread.is_alive()

            ## If no response, calculate greek
            if isThreadRunning:
                print("OptionDataManager calculating greeks. Database unavailable")
                OptDataManagerLogger.info(f"Data for {self.opttick} unavailable for query, proceeding to calculate vol")
                
                    ## Calculating greeks with Close BS IV
                close_greeks_data = self.__calculate_greek(NaData, greek, 'bs_iv')
                NaData[close_greeks_data.columns] = close_greeks_data

                ## Calculating greeks with Midpoint/Weighted_Midpoint BS IV
                if self.default_fill:
                    close_greeks_data = self.__calculate_greek(NaData, greek, major_vol_name)
                    close_greeks_data.columns = [f"{self.default_fill}_{x}" for x in close_greeks_data.columns]
                    NaData[close_greeks_data.columns] = close_greeks_data

                ## Concatenating the calculated greeks. Only concantenating when we have to calculate greeks
                final_data = NaData[returnColumns]
                final_data = pd.concat([final_data, nonNaData[returnColumns]])
                # print(f"Final Data: {len(final_data)}")
                # print(final_data)
                
            else:
                ## If we query database, we do not need to concat. The database has everything
                print('Good to query database', end = '\r')
                if timeAgg == 'INTRA':
                    table = 'temp_options_intra'
                elif timeAgg == 'EOD':
                    table = 'temp_options_eod'
                else:
                    raise ValueError('Unknown timeAgg arguement. Recieved', timeAgg)
                final_data = query_database('securities_master', table, self.query)
                final_data.columns = [x.lower() for x in final_data.columns]
                final_data = final_data[returnColumns]
        else:
            print("Using available data", end = '\r')
            final_data = data[returnColumns]


        final_data.columns = [x.capitalize() for x in final_data.columns]
        final_data.set_index('Datetime', inplace = True)
        # print(final_data)
        return final_data
        


        
    def __calculate_greek(self, unprocessed_data, greeks:list, vol_name):
        greek_function = {
            'delta': Calculate.delta,
            'gamma': Calculate.gamma,
            'theta': Calculate.theta,
            'rho': Calculate.rho,
            'vega': Calculate.vega,
            'vanna': Calculate.vanna,
            'volga': Calculate.volga
        }
        
        Tpool = Pool(5)
        
        results = Tpool.map(
            calculate_single_greeks,
            greeks, 
             [unprocessed_data] * len(greeks), 
             [vol_name] * len(greeks), 
             [greek_function] * len(greeks)
        )
        Tpool.close()
        Tpool.join()
        Tpool.clear()
        Tpool.restart()
        results = pd.concat(results, axis = 1)
        return results

def calculate_single_greeks( greek, unprocessed_data, vol_name, greek_function):
    def safe_calculate(greek_func, **kwargs):
        try:
            return greek_func(**kwargs)
        except:
            return 0.0
    series = unprocessed_data.apply(
        lambda x: safe_calculate(
            greek_function[greek],
            flag=x['put/call'],
            K=x['strike'],
            S=x['underlier_price'],
            exp=x['expiration'],
            r=x['rf_rate'],
            sigma=x[vol_name],
            start=x['datetime'],
            y=x['dividend']
        ),
        axis=1
    )
    series.name = greek
    return series

